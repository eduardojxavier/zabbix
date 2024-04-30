from zabbix_api import ZabbixAPI
import datetime
import time
from dateutil.relativedelta import relativedelta
import json
import pytz
import openpyxl
import pandas as pd

URL = ''
USERNAME = ''
PASSWORD = ''

try:
    zapi = ZabbixAPI(URL, timeout=180)
    zapi.login(USERNAME,PASSWORD)
    print('Conectado com sucesso!')

except Exception as erro:
    print(f'Não foi possível conectar-se. {erro}')
    exit()



lista_eventos = [] 
lista_hosts_em_manutencao = []
lista_final = [] 
lista_tempo = []
lista_final_sem_manutencao = []


def converterData():

    print("Data Inicial do Relatório")
    diaIni = int(input("Entre com o dia: "))
    mesIni = int(input("Entre com o mês: "))
    anoIni = int(input("Entre com o ano: "))

    criaDataInicio = datetime.datetime(anoIni, mesIni, diaIni, 0,1)
    dataInicio = time.mktime(criaDataInicio.timetuple())

    print("Data Final do Relatório")
    diaFim = int(input("Entre com o dia: "))
    mesFim = int(input("Entre com o mês: "))
    anoFim = int(input("Entre com o ano: "))

    criaDataFinal = datetime.datetime(anoFim, mesFim, diaFim, 23, 59)
    dataFinal = time.mktime(criaDataFinal.timetuple())

    print("Executando relatório")
    get_history(lista_eventos, lista_final, dataInicio, dataFinal, lista_tempo, lista_hosts_em_manutencao, lista_final_sem_manutencao)



def get_history(lista_eventos, lista_final, dataInicio, dataFinal, lista_tempo, lista_hosts_em_manutencao, lista_final_sem_manutencao): 

    
    fuso_horario_recife = pytz.timezone('America/Recife')

    abertura_evento = zapi.event.get({'output':['eventid','name','host','severity','clock','r_eventid','acknowledged'],
                                    'selectHosts':['hostid','host'],
                                    'time_from': dataInicio,
                                    'time_till': dataFinal,
                                     })   

         
    #converter o json em uma lista de dicionários
    converter_json = json.dumps(abertura_evento)
    dicionario_abertura_evento = json.loads(converter_json)

    #Testar se o evento está com r_eventid preenchido
    for i in dicionario_abertura_evento:

        #Atribui cada posição do dicionário a uma variável
        id_inicio_evento = i['eventid']
        id_retorno_evento = i['r_eventid']
        nome_evento = i['name']
        severidade_evento = i['severity']
        host_evento = i['hosts'][0]['host']
        id_host_evento = i['hosts'][0]['hostid']
        eventoReconhecido = i['acknowledged']


        if (id_retorno_evento != '0') and (severidade_evento == '4' or severidade_evento == '5'): 
            lista_eventos.append(i)
           
            if severidade_evento == "4":
                severidade = 'Alta'
            elif severidade_evento == "5":
                severidade = 'Desastre'

            if eventoReconhecido == '0':
                reconhecido = 'Não'
            else:
                reconhecido = 'Sim'
            

            for j in lista_eventos:
                converte_hora = int(j['clock'])
                data_hora = datetime.datetime.fromtimestamp(converte_hora,fuso_horario_recife)
                horaInicio = data_hora.strftime("%Y-%m-%d %H:%M:%S")
                      
                
                fechamento_evento = zapi.event.get({"output":['clock'],
                                                    "eventids": id_retorno_evento})
                
                converter_json_fechamento = json.dumps(fechamento_evento)
                dicionario_fechamento_evento = json.loads(converter_json_fechamento)
                for k in dicionario_fechamento_evento:
                    converte_hora_fechamento = int(k['clock'])
                    data_hora_fechamento = datetime.datetime.fromtimestamp(converte_hora_fechamento, fuso_horario_recife)
                    horaFinal = data_hora_fechamento.strftime("%Y-%m-%d %H:%M:%S")

            #Converter por extenso o tempo
            mascara = "%Y-%m-%d %H:%M:%S"
            ini = datetime.datetime.strptime(horaInicio, mascara)
            fim = datetime.datetime.strptime(horaFinal, mascara)
            di = abs(relativedelta(ini, fim))
            tempoEvento = (f'{di.days} dias, {di.hours} horas, {di.minutes} minutos')

            resultado = id_inicio_evento, nome_evento, id_host_evento, host_evento, severidade, horaInicio, horaFinal, tempoEvento, reconhecido
            
            lista_final.append(resultado)

    
    return lista_final
    

        
def indisponibilidade_total(lista_final):
    for evento in lista_final:
        tempo_indisponibilidade = evento[7]
        dias, horas, minutos = tempo_indisponibilidade.split(', ')
        dias = int(dias.split()[0])
        horas = int(horas.split()[0])
        minutos = int(minutos.split()[0])

        total_horas_indisponibilidade = (dias * 24 * 60) + (horas * 60) + minutos 

        lista_tempo.append(total_horas_indisponibilidade)
   
    tempo_total = sum(lista_tempo)

    # indis = str(tempo_total) + " minutos"

    return f'{tempo_total} minutos'


def exportar_para_excel(lista_final, lista_tempo):
    colunas = ['ID Evento', 'Nome do Evento', 'ID Host', 'Nome do Host', 'Severidade', 'Hora inicial', 'Hora final', 'Tempo de indisponibilidade', 'Reconhecido']
    df = pd.DataFrame(lista_final, columns=colunas)

    nome_arquivo_excel = 'resultados_zabbix.xlsx'
    df.to_excel(nome_arquivo_excel, index=False)

    #abrir o arquivo para adicionar a indisponibilidade total
    wb = openpyxl.load_workbook(nome_arquivo_excel)
    ws = wb.active

    # total_indisponibilidade = indisponibilidade_total(lista_final)
    total_indisponibilidade = sum(lista_tempo)
    max_row = ws.max_row + 2
    ws[f'A{max_row}'] = "Indisponibilidade Total"
    ws[f'B{max_row}'] = f'{total_indisponibilidade} minutos'

    wb.save(nome_arquivo_excel)

    print(f'Arquivo {nome_arquivo_excel} gerado com sucesso!')




if __name__ == "__main__":

    converterData()
    print(indisponibilidade_total(lista_final))
    exportar_para_excel(lista_final, lista_tempo)
